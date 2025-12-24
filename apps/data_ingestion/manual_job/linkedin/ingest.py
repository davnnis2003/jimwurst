import os
import sys
import csv
import psycopg2
from psycopg2 import sql
from psycopg2.extras import execute_values
from tqdm import tqdm
from openpyxl import load_workbook

# Optional dependency for richer Excel parsing; fallback to openpyxl if missing.
try:
    import pandas as pd  # type: ignore
    HAS_PANDAS = True
except Exception:
    HAS_PANDAS = False

# Import common utilities
sys.path.append(os.path.join(os.path.dirname(__file__), '../../../../utils'))
from ingestion_utils import load_env, get_db_connection, ensure_schema, clean_header, sanitize_table_name

# Load env vars
load_env()

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_NAME = os.getenv("POSTGRES_DB", "jimwurst_db")
DB_USER = os.getenv("POSTGRES_USER", "jimwurst_user")
DB_PASS = os.getenv("POSTGRES_PASSWORD", "jimwurst_password")
DB_PORT = os.getenv("DB_PORT", "5432")

DEFAULT_DATA_PATH = os.path.expanduser("~/Documents/jimwurst_local_data/linkedin")
DATA_PATH = os.getenv("LINKEDIN_DATA_PATH", DEFAULT_DATA_PATH)
BASIC_PATH = os.path.join(DATA_PATH, "basic")
COMPLETE_PATH = os.path.join(DATA_PATH, "complete")

SCHEMA_NAME = "s_linkedin"


def dedupe_columns(columns):
    """Ensure column names are unique by appending suffixes to duplicates."""
    seen = {}
    unique = []
    for col in columns:
        base = col
        if base in seen:
            seen[base] += 1
            unique.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            unique.append(base)
    return unique


class ExcelIngestor:
    """Handles ingestion of .xlsx files from basic creator insights exports"""
    
    def __init__(self, conn):
        self.conn = conn
    
    def ingest(self, file_path, table_name):
        """Ingest Excel file into PostgreSQL"""
        print(f"Processing Excel file: {table_name}...")
        
        try:
            if HAS_PANDAS:
                self._ingest_with_pandas(file_path, table_name)
            else:
                self._ingest_with_openpyxl(file_path, table_name)
        except Exception as e:
            print(f"Error processing Excel file {file_path}: {e}")

    def _ingest_with_pandas(self, file_path, table_name):
        xls = pd.ExcelFile(file_path, engine="openpyxl")
        for sheet_name in xls.sheet_names:
            df_raw = xls.parse(sheet_name, dtype=str, header=None)
            df_raw = df_raw.dropna(how="all")  # drop fully empty rows
            df_raw = df_raw.dropna(axis=1, how="all")  # drop fully empty cols
            if df_raw.empty:
                print(f"Skipping empty sheet: {sheet_name}")
                continue

            # Choose header row: prefer first row with >=2 non-null cells; fallback to first non-empty.
            header_idx = None
            for idx, row in df_raw.iterrows():
                non_null = row.dropna()
                if len(non_null) >= 2:
                    header_idx = idx
                    break
            if header_idx is None:
                header_idx = df_raw.first_valid_index()
            if header_idx is None:
                print(f"Skipping sheet with no headers: {sheet_name}")
                continue

            header_row = df_raw.loc[header_idx]
            columns = [
                clean_header(str(h)) if (h is not None and str(h).strip() != "") else f"column_{i}"
                for i, h in enumerate(header_row.tolist())
            ]
            columns = dedupe_columns(columns)

            df = df_raw.loc[header_idx + 1 :].copy()
            df.columns = columns
            df = df.dropna(how="all")

            if df.empty:
                print(f"Skipping empty data after header in sheet: {sheet_name}")
                continue

            current_table_name = f"{table_name}_{sanitize_table_name(sheet_name)}" if len(xls.sheet_names) > 1 else table_name
            self._write_dataframe(df, current_table_name)
            print(f"  ✓ Ingested sheet '{sheet_name}' into table '{current_table_name}'")

    def _write_dataframe(self, df, current_table_name):
        cols_def = ", ".join([f"{sql.Identifier(c).as_string(self.conn)} TEXT" for c in df.columns])
        create_query = sql.SQL("CREATE TABLE IF NOT EXISTS {}.{} ({})").format(
            sql.Identifier(SCHEMA_NAME),
            sql.Identifier(current_table_name),
            sql.SQL(cols_def)
        )
        drop_query = sql.SQL("DROP TABLE IF EXISTS {}.{} CASCADE").format(
            sql.Identifier(SCHEMA_NAME),
            sql.Identifier(current_table_name)
        )
        insert_query = sql.SQL("INSERT INTO {}.{} VALUES %s").format(
            sql.Identifier(SCHEMA_NAME),
            sql.Identifier(current_table_name)
        )

        with self.conn.cursor() as cur:
            cur.execute(drop_query)
            cur.execute(create_query)
        batch = df.where(pd.notnull(df), None).values.tolist()
        if batch:
            with self.conn.cursor() as cur:
                execute_values(cur, insert_query, batch, page_size=1000)
            self.conn.commit()

    def _ingest_with_openpyxl(self, file_path, table_name):
        workbook = load_workbook(file_path, data_only=True)

        # Process each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Create table name with sheet suffix if multiple sheets
            if len(workbook.sheetnames) > 1:
                current_table_name = f"{table_name}_{sanitize_table_name(sheet_name)}"
            else:
                current_table_name = table_name

            # Collect rows
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                print(f"Skipping empty sheet: {sheet_name}")
                continue

            def first_header_idx(rows):
                for idx, row in enumerate(rows):
                    if not row:
                        continue
                    non_null = [c for c in row if c is not None and str(c).strip() != ""]
                    if len(non_null) >= 2:
                        return idx
                # fallback: first non-empty row
                for idx, row in enumerate(rows):
                    if row and any(cell is not None for cell in row):
                        return idx
                return None

            header_idx = first_header_idx(rows)
            if header_idx is None:
                print(f"Skipping sheet with no headers: {sheet_name}")
                continue

            headers = rows[header_idx]
            columns = [
                clean_header(str(h)) if h is not None else f"column_{i}"
                for i, h in enumerate(headers)
            ]
            columns = dedupe_columns(columns)

            # Create table
            cols_def = ", ".join([f"{sql.Identifier(c).as_string(self.conn)} TEXT" for c in columns])

            create_query = sql.SQL("CREATE TABLE IF NOT EXISTS {}.{} ({})").format(
                sql.Identifier(SCHEMA_NAME),
                sql.Identifier(current_table_name),
                sql.SQL(cols_def)
            )

            # Drop and recreate for idempotency
            drop_query = sql.SQL("DROP TABLE IF EXISTS {}.{} CASCADE").format(
                sql.Identifier(SCHEMA_NAME),
                sql.Identifier(current_table_name)
            )

            with self.conn.cursor() as cur:
                cur.execute(drop_query)
                cur.execute(create_query)

            # Insert data (rows after header). Skip rows that are entirely empty.
            insert_query = sql.SQL("INSERT INTO {}.{} VALUES %s").format(
                sql.Identifier(SCHEMA_NAME),
                sql.Identifier(current_table_name)
            )

            data_rows = []
            batch_size = 1000

            for row in rows[header_idx + 1:]:
                if not row or all(val is None for val in row):
                    continue

                processed_row = [str(val) if val is not None else None for val in row]

                if len(processed_row) < len(columns):
                    processed_row += [None] * (len(columns) - len(processed_row))
                elif len(processed_row) > len(columns):
                    processed_row = processed_row[:len(columns)]

                data_rows.append(processed_row)

                if len(data_rows) >= batch_size:
                    with self.conn.cursor() as cur:
                        execute_values(cur, insert_query, data_rows)
                    self.conn.commit()
                    data_rows = []

            # Insert remaining rows
            if data_rows:
                with self.conn.cursor() as cur:
                    execute_values(cur, insert_query, data_rows)
                self.conn.commit()

            print(f"  ✓ Ingested sheet '{sheet_name}' into table '{current_table_name}'")

        workbook.close()


class CSVIngestor:
    """Handles ingestion of .csv files from full data archive exports"""
    
    def __init__(self, conn):
        self.conn = conn
    
    def ingest(self, file_path, table_name):
        """Ingest CSV file into PostgreSQL"""
        print(f"Processing CSV file: {table_name}...")
        
        try:
            # Detect encoding - LinkedIn sometimes uses UTF-8 or UTF-16
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                
                if not headers:
                    print(f"Skipping empty file: {file_path}")
                    return

                # Sanitize and dedupe column names
                columns = []
                for i, h in enumerate(headers):
                    if h is None or str(h).strip() == "":
                        columns.append(f"column_{i}")
                    else:
                        columns.append(clean_header(h))
                columns = dedupe_columns(columns)
                
                # Create table
                cols_def = ", ".join([f"{sql.Identifier(c).as_string(self.conn)} TEXT" for c in columns])
                
                create_query = sql.SQL("CREATE TABLE IF NOT EXISTS {}.{} ({})").format(
                    sql.Identifier(SCHEMA_NAME),
                    sql.Identifier(table_name),
                    sql.SQL(cols_def)
                )
                
                # Drop and recreate for idempotency
                drop_query = sql.SQL("DROP TABLE IF EXISTS {}.{} CASCADE").format(
                    sql.Identifier(SCHEMA_NAME),
                    sql.Identifier(table_name)
                )
                
                with self.conn.cursor() as cur:
                    cur.execute(drop_query)
                    cur.execute(create_query)
                
                # Insert data
                insert_query = sql.SQL("INSERT INTO {}.{} VALUES %s").format(
                    sql.Identifier(SCHEMA_NAME),
                    sql.Identifier(table_name)
                )
                
                rows = []
                batch_size = 1000
                
                for row in reader:
                    # Pad or truncate row to match headers
                    if len(row) < len(columns):
                        row += [None] * (len(columns) - len(row))
                    elif len(row) > len(columns):
                        row = row[:len(columns)]
                    
                    rows.append(row)
                    
                    if len(rows) >= batch_size:
                        with self.conn.cursor() as cur:
                            execute_values(cur, insert_query, rows)
                        self.conn.commit()
                        rows = []
                
                if rows:
                    with self.conn.cursor() as cur:
                        execute_values(cur, insert_query, rows)
                    self.conn.commit()
                    
        except Exception as e:
            self.conn.rollback()
            print(f"Error processing CSV file {file_path}: {e}")


def scan_for_files():
    """Scan both basic and complete folders for files"""
    files_to_process = []
    
    # Scan basic folder for .xlsx files
    if os.path.exists(BASIC_PATH):
        for root, dirs, files in os.walk(BASIC_PATH):
            for f in files:
                if f.lower().endswith('.xlsx'):
                    files_to_process.append({
                        'path': os.path.join(root, f),
                        'type': 'excel',
                        'source': 'basic'
                    })
    
    # Scan complete folder for .csv files
    if os.path.exists(COMPLETE_PATH):
        for root, dirs, files in os.walk(COMPLETE_PATH):
            for f in files:
                if f.lower().endswith('.csv'):
                    files_to_process.append({
                        'path': os.path.join(root, f),
                        'type': 'csv',
                        'source': 'complete'
                    })
    
    return files_to_process


def main():
    print(f"LinkedIn Data Ingestion")
    print(f"=" * 50)
    print(f"Base Path: {DATA_PATH}")
    print(f"Basic Exports Path: {BASIC_PATH}")
    print(f"Complete Exports Path: {COMPLETE_PATH}")
    print()
    
    # Create directories if they don't exist
    os.makedirs(BASIC_PATH, exist_ok=True)
    os.makedirs(COMPLETE_PATH, exist_ok=True)
    
    # Scan for files
    files_to_process = scan_for_files()
    
    if not files_to_process:
        print(f"No files found in either:")
        print(f"  - {BASIC_PATH} (.xlsx files)")
        print(f"  - {COMPLETE_PATH} (.csv files)")
        print()
        print("Please export your LinkedIn data and place files in the appropriate folder:")
        print("  • Basic creator insights → basic/")
        print("  • Full data archive → complete/")
        sys.exit(0)

    # Calculate total size and estimate time
    total_size_bytes = sum(os.path.getsize(f['path']) for f in files_to_process)
    total_size_mb = total_size_bytes / (1024 * 1024)
    estimated_seconds = total_size_mb / 5.0
    
    # Group files by type
    excel_files = [f for f in files_to_process if f['type'] == 'excel']
    csv_files = [f for f in files_to_process if f['type'] == 'csv']
    
    print("Found the following LinkedIn files to ingest:")
    print()
    
    if excel_files:
        print(f"📊 Basic Creator Insights (.xlsx) - {len(excel_files)} file(s):")
        for f in excel_files:
            rel_path = os.path.relpath(f['path'], BASIC_PATH)
            print(f"   - {rel_path}")
        print()
    
    if csv_files:
        print(f"📁 Full Data Archive (.csv) - {len(csv_files)} file(s):")
        for f in csv_files:
            rel_path = os.path.relpath(f['path'], COMPLETE_PATH)
            print(f"   - {rel_path}")
        print()
    
    print(f"Total Size: {total_size_mb:.2f} MB")
    print(f"Estimated Processing Time: ~{estimated_seconds:.1f} seconds")
    
    auto_confirm = ('--yes' in sys.argv) or ('-y' in sys.argv)

    if not auto_confirm:
        try:
            response = input("\nDo you want to proceed with LinkedIn data ingestion? (y/N): ").strip().lower()
        except KeyboardInterrupt:
            print("\nOperation cancelled.")
            sys.exit(0)
        
        if response != 'y':
            print("Operation cancelled.")
            sys.exit(0)

    print("\nStarting ingestion...")
    print("=" * 50)
    
    conn = get_db_connection()
    ensure_schema(conn, SCHEMA_NAME)
    
    excel_ingestor = ExcelIngestor(conn)
    csv_ingestor = CSVIngestor(conn)
    
    for file_info in tqdm(files_to_process, desc="Ingesting Files"):
        file_path = file_info['path']
        filename = os.path.basename(file_path)
        table_name = f"{file_info['source']}_{sanitize_table_name(filename)}"
        
        if file_info['type'] == 'excel':
            excel_ingestor.ingest(file_path, table_name)
        elif file_info['type'] == 'csv':
            csv_ingestor.ingest(file_path, table_name)
        
    conn.close()
    print("\n" + "=" * 50)
    print("✅ Ingestion complete!")

if __name__ == "__main__":
    main()
